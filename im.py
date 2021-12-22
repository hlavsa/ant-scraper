from bs4 import BeautifulSoup
from random import randint
from time import sleep
from openpyxl import load_workbook

import requests, json, openpyxl, urllib
import pandas as pd

headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET',
        'Access-Control-Allow-Headers': 'Content-Type',
        'Access-Control-Max-Age': '3600',
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:52.0) Gecko/20100101 Firefox/52.0'
    }

# Load excel, delete NaNs, save excel
df = pd.read_excel('335319.xlsx') 
df.dropna(inplace=True)
df.to_excel('prefinal.xlsx')

# load excel, delete rows n cols, save excel
wb = load_workbook('prefinal.xlsx')
ws = wb.active
ws.delete_cols(5, 1)
ws.delete_cols(1, 2)
ws.delete_rows(1,2)
wb.save('final.xlsx')

# convert excel to txt
with open('test.txt', 'w') as file:
    pd.read_excel('final.xlsx').to_string(file, index=False)


# create final excel template
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = 'seznam antiku'
sheet.append(['název', 'autor', 'odkaz', 'cena', 'antik'])

# main starts here
huge_list = []
with open('test.txt', "r") as f:
  for line in f:
    hg = line.replace('++','').replace('(pseudonym)', '').replace('. ', '').replace('\t', ' ').replace('kolektiv autorů','')
    huge_list.append(hg.strip())


def get_organic_results():
    for query in huge_list:
      query = urllib.parse.quote_plus(query)
      
      #sleep(randint(4,64))      
      html = requests.get('https://www.ulovknihu.cz/hledat?q={}&state%5B%5D=cz&region%5B%5D=1&region%5B%5D=2&region%5B%5D=3&region%5B%5D=4&region%5B%5D=5&region%5B%5D=6&region%5B%5D=7&region%5B%5D=8&region%5B%5D=9&region%5B%5D=10&region%5B%5D=11&region%5B%5D=12&region%5B%5D=13&region%5B%5D=14&stone=0&also_sold=0&sort=2&price_min=&price_max='.format(query), headers=headers, timeout=5).text
      soup = BeautifulSoup(html, 'html.parser')
      
      data = []
      
      soup.select('.my-md-row')
      
      try:
        title = soup.find('div', {'class': 'my-md-td searchList__product__info'}).find('h2').text
      except:
        title = query

      try:
        author = soup.find('div', {'class': 'searchList__product__info__autor'}).find('a').text
      except:
        author = 'Nic nenalezeno'

      try:
        link = soup.find('a', {'class': 'btn searchList__product__vendor__bottom__link'})['href']
      except:
        link = 'Nic nenalezeno'

      try:
        price = float(soup.find('div', {'class': 'searchList__product__vendor__bottom__price'}).text.replace('Kč', '').strip())
      except:
        price = 'Nic nenalezeno'

      try:
        ant_name = soup.find('div', {'class': 'my-md-td searchList__product__vendor'}).find('span').text.replace('\n', '').replace('    ', '').strip()
      except:
        ant_name = 'Nic nenalezeno'

      data.append([title, author, link, price, ant_name])
      sheet.append([title, author, link, price, ant_name])
      
      # pro vic vysledku na jeden dotaz use this:
      # for item in soup.select('.my-md-row'):
      #     title = item.find('div', {'class': 'my-md-td searchList__product__info'}).find('h2').text
      #     link = item.find('a', {'class': 'btn searchList__product__vendor__bottom__link'})['href']

      #     try:
      #         price = float(item.find('div', {'class': 'searchList__product__vendor__bottom__price'}).text.replace('Kč', '').strip())
      #     except:
      #         price = None

      #     data.append({
      #         'item': {'title': title, 'link': link, 'price': price},
      #     })
      #     

      print(data)

get_organic_results()
excel.save('seznam_knih.xlsx')



